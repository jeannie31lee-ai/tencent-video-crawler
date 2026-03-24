#!/usr/bin/env python3
"""生成腾讯视频片库数据Excel V7版本"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

with open("/tmp/tencent_video_v7.json", "r", encoding="utf-8") as f:
    data = json.load(f)

tv = data.get("电视剧", [])
mv = data.get("电影", [])

wb = Workbook()

# ============ Sheet1: 全部详情 ============
ws1 = wb.active
ws1.title = "全部详情"

headers = ["平台", "剧名", "类型", "是否独播", "付费类型", "年份", "演员", "地区", "集数", "题材", "男女频"]
ws1.append(headers)

hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    bottom=Side(style="thin", color="D9D9D9")
)

for cell in ws1[1]:
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for category in ["电视剧", "电影"]:
    for item in data.get(category, []):
        ws1.append([
            item.get("平台", ""), item.get("剧名", ""), item.get("类型", ""),
            item.get("是否独播", ""), item.get("付费类型", ""), item.get("年份", ""),
            item.get("演员", ""), item.get("地区", ""), item.get("集数", ""),
            item.get("题材", ""), item.get("男女频", ""),
        ])

col_widths = {"A": 10, "B": 22, "C": 8, "D": 10, "E": 10, "F": 8, "G": 32, "H": 10, "I": 12, "J": 15, "K": 8}
for col, w in col_widths.items():
    ws1.column_dimensions[col].width = w

ws1.auto_filter.ref = f"A1:K{ws1.max_row}"
ws1.freeze_panes = "A2"

# ============ Sheet2: 汇总统计 ============
ws2 = wb.create_sheet("汇总统计")

sec_font = Font(bold=True, size=12)
tbl_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
tbl_font = Font(bold=True)

def write_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = tbl_font
        c.fill = tbl_fill

ws2["A1"] = "腾讯视频片库数据统计 (V7)"
ws2["A1"].font = Font(bold=True, size=14)
ws2.merge_cells("A1:E1")

ws2["A2"] = "独播定义：国内网络播放平台独播，不含海外平台（Netflix/WeTV）及电视台（央视/卫视）；已排除单集≤20分钟短剧"
ws2["A2"].font = Font(italic=True, size=9, color="666666")
ws2.merge_cells("A2:G2")

ws2["A3"] = "V7修正：电影2016年前独播标记不可信（老电影在1905/优酷等多平台播出），统一改为非独播；电视剧保持API标记+用户修正"
ws2["A3"].font = Font(italic=True, size=9, color="666666")
ws2.merge_cells("A3:G3")

# 一、总体统计
tv_excl = sum(1 for it in tv if it["是否独播"] == "独播")
mv_excl = sum(1 for it in mv if it["是否独播"] == "独播")
tv_total, mv_total = len(tv), len(mv)

r = 5
ws2.cell(r, 1, "一、总体统计").font = sec_font
write_header(ws2, r+1, ["类型", "总数", "独播", "非独播", "独播占比"])

for i, (name, total, excl) in enumerate([(("电视剧", tv_total, tv_excl)), (("电影", mv_total, mv_excl))], r+2):
    ws2.cell(i, 1, name)
    ws2.cell(i, 2, total)
    ws2.cell(i, 3, excl)
    ws2.cell(i, 4, total - excl)
    ws2.cell(i, 5, f"{excl/total*100:.1f}%")

ws2.cell(r+4, 1, "总计").font = Font(bold=True)
ws2.cell(r+4, 2, tv_total + mv_total).font = Font(bold=True)
ws2.cell(r+4, 3, tv_excl + mv_excl)
ws2.cell(r+4, 4, (tv_total - tv_excl) + (mv_total - mv_excl))
pct_total = (tv_excl + mv_excl) / (tv_total + mv_total) * 100
ws2.cell(r+4, 5, f"{pct_total:.1f}%")

# 二、付费类型分布
r = 12
ws2.cell(r, 1, "二、付费类型分布").font = sec_font
write_header(ws2, r+1, ["付费类型", "电视剧", "电影", "总计"])

for i, pt in enumerate(["会员", "免费", "付费"], r+2):
    tc = sum(1 for it in tv if it.get("付费类型") == pt)
    mc = sum(1 for it in mv if it.get("付费类型") == pt)
    ws2.cell(i, 1, pt); ws2.cell(i, 2, tc); ws2.cell(i, 3, mc); ws2.cell(i, 4, tc+mc)

# 三、男女频分布
r = 18
ws2.cell(r, 1, "三、男女频分布").font = sec_font
write_header(ws2, r+1, ["受众", "电视剧", "电影", "总计"])

for i, g in enumerate(["男频", "女频", "通用"], r+2):
    tc = sum(1 for it in tv if it.get("男女频") == g)
    mc = sum(1 for it in mv if it.get("男女频") == g)
    ws2.cell(i, 1, g); ws2.cell(i, 2, tc); ws2.cell(i, 3, mc); ws2.cell(i, 4, tc+mc)

# 四、地区分布
r = 24
ws2.cell(r, 1, "四、地区分布（Top 15）").font = sec_font
write_header(ws2, r+1, ["地区", "电视剧", "电影", "总计"])

area_counts = {}
for items, cat in [(tv, "电视剧"), (mv, "电影")]:
    for it in items:
        a = it.get("地区", "未知") or "未知"
        area_counts.setdefault(a, {"电视剧": 0, "电影": 0})[cat] += 1

for i, (area, counts) in enumerate(sorted(area_counts.items(), key=lambda x: -(x[1]["电视剧"]+x[1]["电影"]))[:15], r+2):
    ws2.cell(i, 1, area); ws2.cell(i, 2, counts["电视剧"]); ws2.cell(i, 3, counts["电影"]); ws2.cell(i, 4, counts["电视剧"]+counts["电影"])

# 五、年份分布
r = 42
ws2.cell(r, 1, "五、年份分布（近10年）").font = sec_font
write_header(ws2, r+1, ["年份", "电视剧", "电影", "总计", "电视剧独播", "电影独播"])

year_stats = {}
for items, cat in [(tv, "电视剧"), (mv, "电影")]:
    for it in items:
        y = str(it.get("年份", "未知"))
        year_stats.setdefault(y, {"电视剧": 0, "电影": 0, "电视剧独播": 0, "电影独播": 0})
        year_stats[y][cat] += 1
        if it["是否独播"] == "独播":
            year_stats[y][f"{cat}独播"] += 1

recent = sorted([y for y in year_stats if y.isdigit() and int(y) >= 2016], reverse=True)
for i, y in enumerate(recent[:11], r+2):
    s = year_stats[y]
    ws2.cell(i, 1, y); ws2.cell(i, 2, s["电视剧"]); ws2.cell(i, 3, s["电影"])
    ws2.cell(i, 4, s["电视剧"]+s["电影"]); ws2.cell(i, 5, s["电视剧独播"]); ws2.cell(i, 6, s["电影独播"])

# 六、独播内容年份分布对比
r = 56
ws2.cell(r, 1, "六、独播电影年份分布").font = sec_font
write_header(ws2, r+1, ["年份段", "独播电影数", "电影总数", "独播占比"])

year_ranges = [
    ("2025", lambda y: y == 2025),
    ("2024", lambda y: y == 2024),
    ("2023", lambda y: y == 2023),
    ("2022", lambda y: y == 2022),
    ("2021", lambda y: y == 2021),
    ("2020", lambda y: y == 2020),
    ("2019", lambda y: y == 2019),
    ("2016-2018", lambda y: 2016 <= y <= 2018),
    ("2015及更早", lambda y: y <= 2015),
]

for i, (label, cond) in enumerate(year_ranges, r+2):
    excl_count = 0
    total_count = 0
    for it in mv:
        y_str = str(it.get("年份", ""))
        try:
            y = int(y_str) if y_str.isdigit() else 0
        except:
            y = 0
        if y > 0 and cond(y):
            total_count += 1
            if it["是否独播"] == "独播":
                excl_count += 1
    ws2.cell(i, 1, label)
    ws2.cell(i, 2, excl_count)
    ws2.cell(i, 3, total_count)
    ws2.cell(i, 4, f"{excl_count/total_count*100:.1f}%" if total_count > 0 else "0%")

for col in ["A","B","C","D","E","F","G"]:
    ws2.column_dimensions[col].width = 15

import os
output_dir = "/Users/liwenjing/.qoderwork/workspace/mmyjwdi9xszlan59/outputs"
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "腾讯视频片库数据_v7.xlsx")
wb.save(output_path)

import shutil
dest = "/Users/liwenjing/Documents/AI/QODER/腾讯视频片库数据_v7.xlsx"
shutil.copy2(output_path, dest)

print(f"Excel saved: {output_path}")
print(f"Also saved to: {dest}")
print(f"电视剧: {tv_total}部 (独播{tv_excl}, 占{tv_excl/tv_total*100:.1f}%)")
print(f"电影: {mv_total}部 (独播{mv_excl}, 占{mv_excl/mv_total*100:.1f}%)")
print(f"总计: {tv_total+mv_total}部")
