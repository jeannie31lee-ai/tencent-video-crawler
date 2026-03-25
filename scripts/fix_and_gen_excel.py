#!/usr/bin/env python3
"""
V8独播修正 + Excel生成
规则：
1. 电视剧：保持API tag_2标记 + 用户累计修正
2. 电影：2016年前独播标记不可信，改为非独播 + 用户修正
"""

import json, os, shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

# ===== 用户累计修正名单 =====
USER_CONFIRMED_EXCLUSIVE_TV = {
    "陈情令", "庆余年", "听雪楼", "梦回", "双世宠妃", "双世宠妃2",
    "三生三世枕上书", "龙岭迷窟", "传闻中的陈芊芊", "有翡",
    "斗罗大陆", "长歌行", "你是我的荣耀", "雪中悍刀行", "司藤", "锦心似玉",
    "开端", "梦华录", "星汉灿烂", "卿卿日常", "猎罪图鉴", "且试天下",
    "长相思", "莲花楼", "狂飙", "三体", "长月烬明",
    "庆余年第二季", "玫瑰的故事", "永夜星河", "九重紫",
    "大奉打更人", "国色芳华", "五福临门", "了不起的曹萱萱",
    "今夕何夕", "我的小确幸", "画江湖之天罡", "我的父亲母亲",
    "她的盛焰", "隐身的名字", "玫瑰丛生", "季雨倾城", "明珠奇谭", "临暗",
    "芳华里", "又野又烈", "藏匿爱意", "枕红妆", "重返青春", "去听旷野的风",
    "医馆笑传", "渗透", "北平无战事", "战火四千金", "连环套",
    "大丈夫", "妈祖", "一仆二主", "神秘人质", "红色", "英雄使命",
}
USER_CONFIRMED_NON_EXCLUSIVE_TV = {
    "香蜜沉沉烬如霜", "三生三世十里桃花",
    "与凤行", "天下正道", "我家住在高岗上", "生死翻盘",
    "毒刺", "铁血雄心",
    "姐妹情缘",
}
USER_CONFIRMED_NON_EXCLUSIVE_MOVIE = {
    "金山", "走路上学", "泥鳅也是鱼", "北川重生", "黄沙渡",
}

# ===== 加载数据 =====
with open("/tmp/tencent_video_v8_raw.json", "r", encoding="utf-8") as f:
    data = json.load(f)

tv = data["电视剧"]
mv = data["电影"]

# ===== 修正独播标识 =====
print("=" * 60)
print("V8 独播修正")
print("=" * 60)

# 电视剧修正
tv_changes = {"to_excl": 0, "to_non": 0}
for it in tv:
    title = it["剧名"]
    old = it["是否独播"]
    if title in USER_CONFIRMED_EXCLUSIVE_TV:
        it["是否独播"] = "独播"
    elif title in USER_CONFIRMED_NON_EXCLUSIVE_TV:
        it["是否独播"] = "非独播"
    if old != it["是否独播"]:
        if it["是否独播"] == "独播": tv_changes["to_excl"] += 1
        else: tv_changes["to_non"] += 1

tv_excl = sum(1 for it in tv if it["是否独播"] == "独播")
print(f"\n电视剧: {len(tv)}部, 独播{tv_excl}部({tv_excl/len(tv)*100:.1f}%)")
print(f"  修正: +{tv_changes['to_excl']}独播, -{tv_changes['to_non']}独播")

# 电影修正
mv_changes = {"to_non_year": 0, "to_non_user": 0}
for it in mv:
    title = it["剧名"]
    old = it["是否独播"]
    year_str = str(it.get("年份", ""))
    try: year = int(year_str) if year_str.isdigit() else 0
    except: year = 0

    if title in USER_CONFIRMED_NON_EXCLUSIVE_MOVIE:
        it["是否独播"] = "非独播"
        if old == "独播": mv_changes["to_non_user"] += 1
    elif year > 0 and year < 2016 and old == "独播":
        it["是否独播"] = "非独播"
        mv_changes["to_non_year"] += 1

mv_excl = sum(1 for it in mv if it["是否独播"] == "独播")
print(f"\n电影: {len(mv)}部, 独播{mv_excl}部({mv_excl/len(mv)*100:.1f}%)")
print(f"  2016前改非独播: {mv_changes['to_non_year']}部, 用户修正: {mv_changes['to_non_user']}部")

total = len(tv) + len(mv)
total_excl = tv_excl + mv_excl
print(f"\n总计: {total}部, 独播{total_excl}部({total_excl/total*100:.1f}%)")

# 保存修正后数据
with open("/tmp/tencent_video_v8.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

# ===== 生成Excel =====
print(f"\n{'=' * 60}")
print("生成Excel")
print("=" * 60)

wb = Workbook()

# Sheet1: 全部详情
ws1 = wb.active
ws1.title = "全部详情"
headers = ["平台", "剧名", "类型", "是否独播", "付费类型", "年份", "演员", "地区", "集数", "题材", "男女频"]
ws1.append(headers)

hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
for cell in ws1[1]:
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for cat in ["电视剧", "电影"]:
    for item in data[cat]:
        ws1.append([item.get(h, "") for h in headers])

col_widths = {"A": 10, "B": 22, "C": 8, "D": 10, "E": 10, "F": 8, "G": 32, "H": 10, "I": 12, "J": 15, "K": 8}
for col, w in col_widths.items():
    ws1.column_dimensions[col].width = w
ws1.auto_filter.ref = f"A1:K{ws1.max_row}"
ws1.freeze_panes = "A2"

# Sheet2: 汇总统计
ws2 = wb.create_sheet("汇总统计")
sec_font = Font(bold=True, size=12)
tbl_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
tbl_font = Font(bold=True)

def write_header(ws, row, hdrs):
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = tbl_font
        c.fill = tbl_fill

ws2["A1"] = "腾讯视频片库数据统计 (V8)"
ws2["A1"].font = Font(bold=True, size=14)
ws2.merge_cells("A1:E1")
ws2["A2"] = "独播定义：国内网络播放平台独播，不含海外平台及电视台；已排除单集≤20分钟短剧；电影2016年前统一按非独播处理"
ws2["A2"].font = Font(italic=True, size=9, color="666666")
ws2.merge_cells("A2:G2")

# 一、总体统计
r = 4
ws2.cell(r, 1, "一、总体统计").font = sec_font
write_header(ws2, r+1, ["类型", "总数", "独播", "非独播", "独播占比"])
for i, (name, items) in enumerate([(("电视剧", tv)), (("电影", mv))], r+2):
    excl = sum(1 for it in items if it["是否独播"] == "独播")
    ws2.cell(i, 1, name); ws2.cell(i, 2, len(items)); ws2.cell(i, 3, excl)
    ws2.cell(i, 4, len(items)-excl); ws2.cell(i, 5, f"{excl/len(items)*100:.1f}%")
ws2.cell(r+4, 1, "总计").font = Font(bold=True)
ws2.cell(r+4, 2, total).font = Font(bold=True)
ws2.cell(r+4, 3, total_excl); ws2.cell(r+4, 4, total-total_excl)
ws2.cell(r+4, 5, f"{total_excl/total*100:.1f}%")

# 二、付费类型分布
r = 10
ws2.cell(r, 1, "二、付费类型分布").font = sec_font
write_header(ws2, r+1, ["付费类型", "电视剧", "电影", "总计"])
for i, pt in enumerate(["会员", "免费", "付费"], r+2):
    tc = sum(1 for it in tv if it.get("付费类型") == pt)
    mc = sum(1 for it in mv if it.get("付费类型") == pt)
    ws2.cell(i, 1, pt); ws2.cell(i, 2, tc); ws2.cell(i, 3, mc); ws2.cell(i, 4, tc+mc)

# 三、男女频分布
r = 16
ws2.cell(r, 1, "三、男女频分布").font = sec_font
write_header(ws2, r+1, ["受众", "电视剧", "电影", "总计"])
for i, g in enumerate(["男频", "女频", "通用"], r+2):
    tc = sum(1 for it in tv if it.get("男女频") == g)
    mc = sum(1 for it in mv if it.get("男女频") == g)
    ws2.cell(i, 1, g); ws2.cell(i, 2, tc); ws2.cell(i, 3, mc); ws2.cell(i, 4, tc+mc)

# 四、地区分布
r = 22
ws2.cell(r, 1, "四、地区分布（Top 15）").font = sec_font
write_header(ws2, r+1, ["地区", "电视剧", "电影", "总计"])
area_counts = {}
for items, cat in [(tv, "电视剧"), (mv, "电影")]:
    for it in items:
        a = it.get("地区", "未知") or "未知"
        area_counts.setdefault(a, {"电视剧": 0, "电影": 0})[cat] += 1
for i, (area, counts) in enumerate(sorted(area_counts.items(), key=lambda x: -(x[1]["电视剧"]+x[1]["电影"]))[:15], r+2):
    ws2.cell(i, 1, area); ws2.cell(i, 2, counts["电视剧"]); ws2.cell(i, 3, counts["电影"]); ws2.cell(i, 4, counts["电视剧"]+counts["电影"])

# 五、年份分布
r = 40
ws2.cell(r, 1, "五、年份分布（近10年）").font = sec_font
write_header(ws2, r+1, ["年份", "电视剧", "电影", "总计", "电视剧独播", "电影独播"])
year_stats = {}
for items, cat in [(tv, "电视剧"), (mv, "电影")]:
    for it in items:
        y = str(it.get("年份", "未知"))
        year_stats.setdefault(y, {"电视剧": 0, "电影": 0, "电视剧独播": 0, "电影独播": 0})
        year_stats[y][cat] += 1
        if it["是否独播"] == "独播":
            year_stats[y][f"{cat}独播"] += 1
recent = sorted([y for y in year_stats if y.isdigit() and int(y) >= 2016], reverse=True)
for i, y in enumerate(recent[:11], r+2):
    s = year_stats[y]
    ws2.cell(i, 1, y); ws2.cell(i, 2, s["电视剧"]); ws2.cell(i, 3, s["电影"])
    ws2.cell(i, 4, s["电视剧"]+s["电影"]); ws2.cell(i, 5, s["电视剧独播"]); ws2.cell(i, 6, s["电影独播"])

for col in ["A","B","C","D","E","F","G"]:
    ws2.column_dimensions[col].width = 15

# 保存
output_dir = "/Users/liwenjing/.qoderwork/workspace/mmyjwdi9xszlan59/outputs"
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "腾讯视频片库数据_v8.xlsx")
wb.save(output_path)

dest = "/Users/liwenjing/Documents/AI/QODER/腾讯视频片库数据_v8.xlsx"
shutil.copy2(output_path, dest)

print(f"\nExcel saved: {dest}")
print(f"电视剧: {len(tv)}部 (独播{tv_excl}, 占{tv_excl/len(tv)*100:.1f}%)")
print(f"电影: {len(mv)}部 (独播{mv_excl}, 占{mv_excl/len(mv)*100:.1f}%)")
print(f"总计: {total}部")
