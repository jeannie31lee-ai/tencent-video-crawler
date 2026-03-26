#!/usr/bin/env python3
"""
V10: Apply comprehensive filters and generate Excel
  - Rule 1: 预告片/外站 (positive_trailer=2 from page scan)
  - Rule 2: 横屏短剧 (content_from=8375890, 仅电视剧)
  - Rule 3: 已下架 (video removed from page scan)
  - Rule 4: 宣传片/非正式内容 (特定标题排除)
  - Rule 5: 竖屏短剧/微短剧 (page scan 竖屏/微短剧标记)
  - Rule 6: 幕后纪录片 (main_genre=纪录片 + 幕后/附属关键词, 仅电影)
  - Rule 7: 无资源 (page scan 暂无资源/无视频)
  - Rule 8: 附属特辑/番外篇 (标题含《xxx》特辑/番外/回顾特辑等)
  - Rule 9: 创意短片/短剧/微电影 (标题含短片/短剧/创意短/微电影)
  - Exclusive corrections: V7 user-confirmed lists + 2016 movie cutoff
"""
import json, re, sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CRAWL_FILE = "/tmp/tencent_v9_crawl.json"
FILTER_FILE = "/tmp/tencent_v9_filter_cids.json"
TRAILER_FILE = "/tmp/v9_trailer_cids.json"
VERTICAL_FILE = "/tmp/v9_vertical_scan.json"
OUTPUT_JSON = "/tmp/tencent_video_v10.json"
OUTPUT_EXCEL = "/Users/liwenjing/.qoderwork/workspace/mmyjwdi9xszlan59/outputs/腾讯视频片单_V10.xlsx"

# ===== V7 独播修正列表 (完整继承) =====
USER_CONFIRMED_EXCLUSIVE_TV = {
    "陈情令", "庆余年", "听雪楼", "梦回", "双世宠妃", "双世宠妃2",
    "三生三世枕上书", "龙岭迷窟", "传闻中的陈芊芊", "有翡",
    "斗罗大陆", "长歌行", "你是我的荣耀", "雪中悍刀行", "司藤", "锦心似玉",
    "开端", "梦华录", "星汉灿烂", "卿卿日常", "猎罪图鉴", "且试天下",
    "长相思", "莲花楼", "狂飙", "三体", "长月烬明",
    "庆余年第二季", "玫瑰的故事", "永夜星河", "九重紫",
    "大奉打更人", "国色芳华", "五福临门", "了不起的曹萱萱",
    "今夕何夕", "我的小确幸", "画江湖之天罡", "我的父亲母亲",
    "她的盛焰", "隐身的名字", "玫瑰丛生", "季雨倾城",
    "重返青春",
    # 2015年前 - 用户明确确认
    "医馆笑传", "渗透", "北平无战事", "战火四千金", "连环套",
    "大丈夫", "妈祖", "一仆二主", "神秘人质", "红色", "英雄使命",
}

USER_CONFIRMED_NON_EXCLUSIVE_TV = {
    "香蜜沉沉烬如霜", "三生三世十里桃花",
    "与凤行", "天下正道", "我家住在高岗上", "生死翻盘",
    "毒刺", "铁血雄心",
    "姐妹情缘",
}

USER_CONFIRMED_NON_EXCLUSIVE_MOVIE = {
    "金山", "走路上学", "泥鳅也是鱼", "北川重生", "黄沙渡",
}

USER_CONFIRMED_EXCLUSIVE_MOVIE = set()

# ===== 宣传片/非正式内容排除 =====
TITLE_BLACKLIST = {
    "繁花剧场版",  # 宣传片，非电视剧或电影
}

# ===== 幕后/附属纪录片关键词 (仅电影) =====
DOC_TITLE_KEYWORDS = ["幕后", "独家纪录", "制作特辑", "拍摄日记", "独家幕后"]
DOC_TAG_KEYWORDS = ["幕后故事", "电影主创采访", "剧组生活"]

# ===== 竖屏短剧手动补充 CID (页面扫描未捕获的) =====
EXTRA_VERTICAL_CIDS = {
    "1wl6xutli1kwyk5",  # 失恋阵线联盟
    "b1kjqvblckgdjkt",  # 换珠格格
}

# ===== 无资源手动补充 CID =====
EXTRA_NO_RESOURCE_CIDS = {
    "xzxaw84pjdlg433",  # 半夜叫你别回头2
}

# ===== 附属特辑/番外篇过滤 (Rule 8) =====
# 匹配 《xxx》+特辑/番外、xxx回顾特辑、独家番外、鹅斯卡X《xxx》、xxx番外篇(结尾)
RE_ATTACHED_SPECIAL = re.compile(
    r'《[^》]+》[^《]*(?:特辑|番外)'   # 《xxx》特辑/番外
    r'|回顾特辑'                        # xxx回顾特辑
    r'|独家番外'                        # xxx独家番外
    r'|鹅斯卡'                          # 鹅斯卡X《xxx》特辑
    r'|番外篇$'                         # xxx番外篇(标题末尾)
)

# ===== 创意短片/短剧/微电影过滤 (Rule 9) =====
RE_SHORT_FILM = re.compile(r'短片|短剧|创意短|微电影')

FEMALE_GENRES = {"爱情", "家庭", "青春", "古装", "宫斗", "甜宠", "都市"}
MALE_GENRES = {"军旅", "刑侦", "竞技", "武侠", "科幻", "战争", "谍战", "悬疑", "权谋", "猎奇"}

def determine_gender(main_genre, genre_tags_str):
    all_genres = set()
    if main_genre: all_genres.add(main_genre)
    if genre_tags_str:
        for g in genre_tags_str.split("、"): all_genres.add(g.strip())
    f = bool(all_genres & FEMALE_GENRES)
    m = bool(all_genres & MALE_GENRES)
    if f and not m: return "女频"
    if m and not f: return "男频"
    return "通用"


# ===== Load data =====
with open(CRAWL_FILE, "r") as f:
    crawl = json.load(f)
with open(FILTER_FILE, "r") as f:
    filters = json.load(f)
with open(TRAILER_FILE, "r") as f:
    trailer_data = json.load(f)
with open(VERTICAL_FILE, "r") as f:
    vertical_data = json.load(f)

# Build filter CID sets
short_tv = set(filters["short_drama_tv_cids"])
short_mv = set(filters["short_drama_mv_cids"])
trailer_tv = set(trailer_data["trailer_tv_cids"])
trailer_mv = set(trailer_data["trailer_mv_cids"])
removed_mv = set(trailer_data["removed_mv_cids"])
vertical_tv = set(vertical_data["vertical_tv_cids"]) | EXTRA_VERTICAL_CIDS
no_resource = set(vertical_data["no_resource_cids"]) | EXTRA_NO_RESOURCE_CIDS

print(f"加载数据: TV={len(crawl['tv'])}, Movie={len(crawl['movie'])}")
print(f"过滤CID: 预告片/外站 TV={len(trailer_tv)}, Movie={len(trailer_mv)}")
print(f"          横屏短剧 TV={len(short_tv)}, Movie={len(short_mv)}")
print(f"          已下架 Movie={len(removed_mv)}")
print(f"          竖屏短剧 TV={len(vertical_tv)}")
print(f"          无资源={len(no_resource)}")

output = {}
stats = {}

for category, items, tr_cids, sd_cids in [
    ("电视剧", crawl["tv"], trailer_tv, short_tv),
    ("电影", crawl["movie"], trailer_mv, short_mv),
]:
    kept = []
    rm_trailer = []   # 预告片/外站/待播
    rm_short = []     # 横屏短剧
    rm_removed = []   # 已下架
    rm_promo = []     # 宣传片/非正式内容
    rm_vertical = []  # 竖屏短剧/微短剧
    rm_doc = []       # 幕后纪录片
    rm_nores = []     # 无资源
    rm_special = []   # 附属特辑/番外篇
    rm_shortfilm = [] # 创意短片/短剧/微电影

    for it in items:
        cid = it["cid"]
        title = it["title"]

        # Rule 4: 宣传片/非正式内容 (标题黑名单)
        if title in TITLE_BLACKLIST:
            rm_promo.append(title)
            continue

        # Rule 1: 预告片/外站 (positive_trailer=2)
        if cid in tr_cids:
            rm_trailer.append(title)
            continue

        # Rule 3: 已下架 (video removed)
        if category == "电影" and cid in removed_mv:
            rm_removed.append(title)
            continue

        # Rule 7: 无资源 (page scan + 手动补充)
        if cid in no_resource:
            rm_nores.append(title)
            continue

        # Rule 8: 附属特辑/番外篇 (标题正则匹配)
        if RE_ATTACHED_SPECIAL.search(title):
            rm_special.append(title)
            continue

        # Rule 9: 创意短片/短剧/微电影 (标题正则匹配)
        if RE_SHORT_FILM.search(title):
            rm_shortfilm.append(title)
            continue

        # Rule 2: 横屏短剧 (content_from=8375890)
        # 仅对电视剧生效（电影中该标记不可靠）
        if category == "电视剧" and cid in sd_cids:
            if title not in USER_CONFIRMED_EXCLUSIVE_TV:
                rm_short.append(title)
                continue

        # Rule 5: 竖屏短剧/微短剧 (page scan 竖屏/微短剧标记 + 手动补充)
        if category == "电视剧" and cid in vertical_tv:
            rm_vertical.append(title)
            continue

        # Rule 6: 幕后纪录片 (仅电影, main_genre=纪录片 + 幕后关键词)
        if category == "电影" and it.get("main_genre") == "纪录片":
            tags = it.get("genre_tags", "")
            is_behind = (any(k in title for k in DOC_TITLE_KEYWORDS)
                        or any(k in tags for k in DOC_TAG_KEYWORDS))
            if is_behind:
                rm_doc.append(title)
                continue

        kept.append(it)

    # ===== 独播修正 =====
    excl_changes = {"to_excl": [], "to_non": []}
    for it in kept:
        title = it["title"]
        old = it["exclusive"]

        if category == "电视剧":
            if title in USER_CONFIRMED_EXCLUSIVE_TV:
                new = "独播"
            elif title in USER_CONFIRMED_NON_EXCLUSIVE_TV:
                new = "非独播"
            else:
                new = old
        else:  # 电影
            if title in USER_CONFIRMED_EXCLUSIVE_MOVIE:
                new = "独播"
            elif title in USER_CONFIRMED_NON_EXCLUSIVE_MOVIE:
                new = "非独播"
            elif old == "独播":
                try:
                    y = int(it["year"]) if it["year"] else 9999
                    if y < 2016:
                        new = "非独播"  # 2016年前电影独播标签不可信
                    else:
                        new = old
                except:
                    new = old
            else:
                new = old

        if old != new:
            if new == "独播":
                excl_changes["to_excl"].append(title)
            else:
                excl_changes["to_non"].append(title)
        it["exclusive"] = new

    print(f"\n{'='*50}")
    print(f"{category}:")
    print(f"  原始: {len(items)}")
    print(f"  移除(预告片/外站): {len(rm_trailer)}")
    print(f"  移除(横屏短剧): {len(rm_short)}")
    print(f"  移除(竖屏短剧): {len(rm_vertical)}")
    print(f"  移除(幕后纪录片): {len(rm_doc)}")
    print(f"  移除(附属特辑/番外): {len(rm_special)}")
    print(f"  移除(短片/短剧/微电影): {len(rm_shortfilm)}")
    print(f"  移除(已下架): {len(rm_removed)}")
    print(f"  移除(无资源): {len(rm_nores)}")
    print(f"  移除(宣传片): {len(rm_promo)}")
    print(f"  独播修正: +{len(excl_changes['to_excl'])} -{len(excl_changes['to_non'])}")
    print(f"  保留: {len(kept)}")
    if rm_trailer: print(f"  [预告片/外站样本]: {rm_trailer[:8]}")
    if rm_short: print(f"  [横屏短剧样本]: {rm_short[:8]}")
    if rm_vertical: print(f"  [竖屏短剧]: {rm_vertical}")
    if rm_doc: print(f"  [幕后纪录片样本]: {rm_doc[:8]}")
    if rm_special: print(f"  [附属特辑/番外]: {rm_special}")
    if rm_shortfilm: print(f"  [短片/短剧/微电影]: {rm_shortfilm}")
    if rm_nores: print(f"  [无资源]: {rm_nores}")
    if rm_promo: print(f"  [宣传片样本]: {rm_promo[:8]}")
    if excl_changes["to_excl"]: print(f"  [新增独播]: {excl_changes['to_excl'][:10]}")
    if excl_changes["to_non"]: print(f"  [移除独播]: 共{len(excl_changes['to_non'])}部")

    output[category] = []
    for it in kept:
        output[category].append({
            "平台": "腾讯视频",
            "剧名": it["title"],
            "类型": category,
            "是否独播": it["exclusive"],
            "付费类型": it["pay_type"],
            "年份": it["year"],
            "演员": it["leading_actor"],
            "地区": it["area_name"],
            "集数": it["episode_info"],
            "题材": it["main_genre"],
            "男女频": determine_gender(it["main_genre"], it.get("genre_tags", "")),
        })
    stats[category] = {
        "total": len(kept),
        "exclusive": sum(1 for r in output[category] if r["是否独播"] == "独播"),
        "non_exclusive": sum(1 for r in output[category] if r["是否独播"] == "非独播"),
        "pay": Counter(r["付费类型"] for r in output[category]),
        "area": Counter(r["地区"] for r in output[category]),
        "genre": Counter(r["题材"] for r in output[category]),
        "year": Counter(r["年份"] for r in output[category]),
        "gender": Counter(r["男女频"] for r in output[category]),
        "removed_trailer": len(rm_trailer),
        "removed_short": len(rm_short),
        "removed_vertical": len(rm_vertical),
        "removed_doc": len(rm_doc),
        "removed_special": len(rm_special),
        "removed_shortfilm": len(rm_shortfilm),
        "removed_removed": len(rm_removed),
        "removed_nores": len(rm_nores),
        "removed_promo": len(rm_promo),
    }

# Save JSON
with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

total = sum(len(v) for v in output.values())
print(f"\n{'='*50}")
print(f"总计: {total} 部")
for cat in ["电视剧", "电影"]:
    s = stats[cat]
    pct = s["exclusive"] / s["total"] * 100 if s["total"] else 0
    print(f"  {cat}: {s['total']}部 (独播{s['exclusive']}部/{pct:.1f}%)")

# ===== 验证关键项目 =====
print(f"\n{'='*50}")
print("验证关键项目:")
v_titles = {
    "应过滤(预告/外站)": ["粗野派", "出不去的房间", "红豆", "去你的岛", "更好的我", "我的爷爷", "云水谣", "极限返航", "八府巡按"],
    "应过滤(宣传片)": ["繁花剧场版"],
    "应过滤(待播)": ["白日提灯", "庆余年第三季", "雪中悍刀行第二季"],
    "应过滤(横屏短剧)": ["万福金安", "沉靡", "明珠奇谭", "芳华里", "临暗", "又野又烈", "藏匿爱意", "枕红妆", "去听旷野的风"],
    "应过滤(竖屏短剧)": ["失恋阵线联盟", "看不见的爸妈", "换珠格格", "口袋恋人第二季"],
    "应过滤(幕后纪录片)": ["《镖人：风起大漠》幕后纪录片", "《扎职2:江湖陌路》独家纪录片", "熊猫计划2幕后纪录"],
    "应过滤(无资源)": ["半夜叫你别回头2"],
    "应过滤(附属特辑)": ["《我们一起摇太阳》独家番外篇", "鹅斯卡X《封神第一部》特辑", "《长相思 第一季》回顾特辑", "一起同过窗第三季番外篇"],
    "应过滤(短片/微电影)": ["聚焦新型腐败和隐性腐败系列创意短剧", "诞辰：首部全虚拟拍摄科幻短片", "21视记微电影计划", "调音师（短片）"],
    "应存在(独播)": ["陈情令", "庆余年", "三体", "隐身的名字", "她的盛焰", "庆余年第二季", "狂飙",
                     "重返青春", "长相思"],
    "应存在(独立作品保留)": ["银河护卫队：假日特辑", "沙海番外之蚌人", "沙海番外之画媒", "西游番外篇之笨妖怪", "老友记重聚特辑"],
}

all_output_titles = {}
for cat in ["电视剧", "电影"]:
    for it in output[cat]:
        all_output_titles.setdefault(it["剧名"], []).append((cat, it["是否独播"]))

errors = 0
for group, titles in v_titles.items():
    for t in titles:
        entries = all_output_titles.get(t, [])
        if entries:
            if "应过滤" in group:
                for cat, excl in entries:
                    print(f"  !! {t}: 仍在输出中 ({cat}, {excl})")
                errors += 1
            elif "独播" in group:
                # 检查是否有任意一条匹配预期（独播）
                found_match = any(excl == "独播" for _, excl in entries)
                for cat, excl in entries:
                    mark = "OK" if excl == "独播" else "  "
                    print(f"  {mark} {t}: {cat} {excl}")
                if not found_match:
                    errors += 1
            else:
                # 应存在(纪录片保留等) - 只检查存在
                for cat, excl in entries:
                    print(f"  OK {t}: {cat} {excl}")
        else:
            if "应存在" in group:
                print(f"  !! {t}: 缺失!")
                errors += 1
            else:
                print(f"  OK {t}: 已过滤")

print(f"\n验证总错误: {errors}")

# ===== Generate Excel =====
wb = Workbook()
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Sheet 1: 全部详情
ws1 = wb.active
ws1.title = "全部详情"
headers = ["平台", "剧名", "类型", "是否独播", "付费类型", "年份", "演员", "地区", "集数", "题材", "男女频"]
col_widths = [12, 25, 8, 10, 10, 8, 40, 10, 12, 10, 8]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER
    ws1.column_dimensions[get_column_letter(col_idx)].width = w

ws1.auto_filter.ref = f"A1:K1"
ws1.freeze_panes = "A2"

row_idx = 2
alt_fill = PatternFill("solid", fgColor="F2F6FC")
for cat in ["电视剧", "电影"]:
    for item in output[cat]:
        for col_idx, h in enumerate(headers, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=item.get(h, ""))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if h == "是否独播" and item[h] == "独播":
                cell.font = Font(color="C00000", bold=True)
        row_idx += 1

# Sheet 2: 汇总统计
ws2 = wb.create_sheet("汇总统计")
SECTION_FONT = Font(bold=True, size=13, color="2F5496")
SUB_FONT = Font(bold=True, size=11)
NUM_FONT = Font(size=11)

def write_section(ws, start_row, title, data_dict, cat_label):
    ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    r = start_row + 1
    for cat in ["电视剧", "电影"]:
        ws.cell(row=r, column=1, value=f"【{cat}】").font = SUB_FONT
        r += 1
        s = stats[cat]
        for k, v in sorted(data_dict(s).items(), key=lambda x: -x[1]):
            ws.cell(row=r, column=2, value=str(k)).font = NUM_FONT
            ws.cell(row=r, column=3, value=v).font = NUM_FONT
            ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
            r += 1
        r += 1
    return r

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 24
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 20
ws2.column_dimensions["E"].width = 12

r = 1
ws2.cell(row=r, column=1, value="腾讯视频片单 V10 汇总统计").font = Font(bold=True, size=16, color="2F5496")
ws2.merge_cells("A1:E1")
r = 3

# Overview
ws2.cell(row=r, column=1, value="数据概览").font = SECTION_FONT
r += 1
for cat in ["电视剧", "电影"]:
    s = stats[cat]
    ws2.cell(row=r, column=1, value=f"【{cat}】").font = SUB_FONT
    r += 1
    for label, val in [
        ("总数", s["total"]),
        ("独播", s["exclusive"]),
        ("非独播", s["non_exclusive"]),
        ("已过滤(预告片/外站)", s["removed_trailer"]),
        ("已过滤(横屏短剧)", s["removed_short"]),
        ("已过滤(竖屏短剧)", s["removed_vertical"]),
        ("已过滤(幕后纪录片)", s["removed_doc"]),
        ("已过滤(附属特辑/番外)", s["removed_special"]),
        ("已过滤(短片/短剧/微电影)", s["removed_shortfilm"]),
        ("已过滤(已下架)", s["removed_removed"]),
        ("已过滤(无资源)", s["removed_nores"]),
        ("已过滤(宣传片)", s["removed_promo"]),
    ]:
        ws2.cell(row=r, column=2, value=label)
        ws2.cell(row=r, column=3, value=val).alignment = Alignment(horizontal="right")
        r += 1
    r += 1

r += 1
r = write_section(ws2, r, "按付费类型", lambda s: dict(s["pay"]), "pay")
r += 1
r = write_section(ws2, r, "按地区", lambda s: dict(s["area"]), "area")
r += 1
r = write_section(ws2, r, "按题材", lambda s: dict(s["genre"]), "genre")
r += 1
r = write_section(ws2, r, "按年份", lambda s: dict(s["year"]), "year")
r += 1
r = write_section(ws2, r, "按男女频", lambda s: dict(s["gender"]), "gender")

wb.save(OUTPUT_EXCEL)
print(f"\nExcel 已保存到: {OUTPUT_EXCEL}")
print(f"JSON 已保存到: {OUTPUT_JSON}")
